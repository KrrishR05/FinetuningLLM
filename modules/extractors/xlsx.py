from __future__ import annotations

import io
from typing import List, Tuple

from modules.cleaner import full_clean
from modules.extractors._base import BaseExtractor
from modules.models import PageContent, Section, TableData


class XLSXExtractor(BaseExtractor):

    def extract(self, data: bytes) -> Tuple[List[PageContent], List[str]]:
        from openpyxl import load_workbook

        pages: List[PageContent] = []
        warnings: List[str] = []

        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            warnings.append(f"Failed to open XLSX: {e}")
            return pages, warnings

        for i, sheet_name in enumerate(wb.sheetnames):
            page_num = i + 1
            sections = [Section(title=f"Sheet: {sheet_name}", level=1, text="", page_number=page_num)]

            try:
                ws = wb[sheet_name]
                lines: list[str] = []
                header_row: list[str] = []
                all_rows: list[list[str]] = []

                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    cell_values = [str(c).strip() if c is not None else "" for c in row]
                    all_rows.append(cell_values)

                    if row_idx == 0:
                        header_row = cell_values
                        lines.append("Columns: " + " | ".join(v for v in cell_values if v))
                        lines.append("")
                        continue

                    pairs = []
                    for j, val in enumerate(cell_values):
                        if val:
                            col_name = header_row[j] if j < len(header_row) and header_row[j] else f"Col{j + 1}"
                            pairs.append(f"{col_name}: {val}")
                    if pairs:
                        lines.append("; ".join(pairs))

                text = full_clean("\n".join(lines))
                if not text.strip():
                    warnings.append(f'Sheet "{sheet_name}" had no data')

                tables: list[TableData] = []
                if all_rows:
                    headers = all_rows[0] if len(all_rows) > 0 else []
                    rows = all_rows[1:] if len(all_rows) > 1 else []
                    tables.append(TableData(
                        headers=headers,
                        rows=rows,
                        page_number=page_num,
                        caption=f"Sheet: {sheet_name}",
                    ))

                pages.append(PageContent(
                    page_number=page_num,
                    text=text,
                    source_label=f"Sheet: {sheet_name}",
                    sections=sections,
                    tables=tables,
                ))
            except Exception as e:
                warnings.append(f'Error reading sheet "{sheet_name}": {e}')
                pages.append(PageContent(
                    page_number=page_num,
                    text="",
                    source_label=f"Sheet: {sheet_name}",
                    sections=sections,
                ))

        wb.close()
        return pages, warnings


def extract_xlsx(data: bytes) -> Tuple[List[PageContent], List[str]]:
    return XLSXExtractor().extract(data)
